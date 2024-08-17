# filemanagement/file_handlers.py
import io
from PyPDF2 import PdfWriter, PdfReader
from openpyxl import Workbook
import csv

def generate_invoice_file(invoice):
    # Example using PyPDF2 to create a simple PDF
    buffer = io.BytesIO()
    pdf_writer = PdfWriter()
    page = pdf_writer.add_blank_page(width=612, height=792)
    page.merge_page(PdfReader(io.BytesIO(f"Invoice #{invoice.id}".encode())).pages[0])
    pdf_writer.write(buffer)
    buffer.seek(0)
    return buffer, 'application/pdf'

def generate_sale_file(sale):
    # Example using openpyxl to create an Excel file
    wb = Workbook()
    ws = wb.active
    ws['A1'] = f"Sale #{sale.id}"
    ws['A2'] = f"Amount: ${sale.amount}"
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

def generate_report_file(report):
    # Example creating a CSV file
    buffer = io.StringIO()
    writer = csv.writer(buffer)
    writer.writerow(['Report ID', 'Date', 'Content'])
    writer.writerow([report.id, report.date, report.content])
    buffer.seek(0)
    return io.BytesIO(buffer.getvalue().encode()), 'text/csv'